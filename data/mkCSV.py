
import openpyxl,os
import numpy as np

files = ['EDDF01000.xlsx','EDDF03000.xlsx','EDDF04000.xlsx']

#data
DI = []
features1 = []
features2 = []
population = []

# load 'EDDF01000.xlsx'
book = openpyxl.load_workbook('EDDF01000.xlsx')
sheet = book["DI先行指数"]

# get elements from 'EDDF01000.xlsx'
for i in range(17,28):
    for j in range(3,15):
        DI.append(float(sheet.cell(row=i, column=j).value))

# get elements from 'EDDF03000.xlsx'
book = openpyxl.load_workbook('EDDF03000.xlsx')
sheet = book['先行系列']
for i in range(126,258):
    for j in [4,5,6,9,11]:
        features1.append(float(sheet.cell(row=i, column=j).value))
features1_add =[]
sheet = book['一致系列']
for i in range(126,258):
    for j in [4,6,7,8]:
        features1_add.append(float(sheet.cell(row=i, column=j).value))

features1_add = np.array(features1_add).reshape(-1,4)


# get elements from 'EDDF04000.xlsx'
book = openpyxl.load_workbook('EDDF04000.xlsx')
sheet = book['ＣＩ先行指数']
for i in range(12,23):
    for j in range(3,15):
        features2.append(float(sheet.cell(row=i, column=j).value))

Di = np.array(DI).reshape(-1,1)
features1 = np.array(features1).reshape(-1,5)
features1 = np.concatenate([features1,features1_add],1)
features2 = np.array(features2).reshape(-1,1)
# input matrix
X = np.concatenate([features1,features2],1)
X = np.concatenate([X,Di],1)
print(X.shape)


# save X
if "sample.xlsx" not in os.listdir(os.getcwd()):
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(132):
        for j in range(11):
            ws.cell(i+1, j+1, value=X[i,j])
    wb.save("sample.xlsx")
else:
    print('Already exsists!')

print('Successfully done!')

